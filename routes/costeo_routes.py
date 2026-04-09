from flask import Blueprint, render_template, request, jsonify
from utils.auth import login_requerido, permiso_modulo
from utils.sheet_cache import obtener_datos
from utils.costeo_manager import cargar_reglas, guardar_mapeo, guardar_costo_directo, guardar_regla_gasto, obtener_costos_efectivos, guardar_prorrateo_adm, obtener_prorrateo_adm
import pandas as pd
from datetime import datetime
from routes.contab_routes import calcular_matriz_gestion, cargar_prorrateos
import io
from flask import send_file
from openpyxl import load_workbook
from utils.utils_excel import aplicar_formato_numerico_excel

costeo_bp = Blueprint("costeo", __name__, url_prefix="/costeo")

def obtener_alias_sucursal(sucursal):
    s = str(sucursal).upper().strip().replace("Ñ", "N")
    alias = [s]
    if "FOOD" in s or "ESC" in s or "MILITAR" in s:
        alias.extend(["ESC. MILITAR", "ESCUELA MILITAR", "FOOD TRUCK", "FOODTRUCK"])
    if "WEB" in s:
        alias.extend(["WEB", "PAGINA WEB", "PAGINAWEB"])
    if "COSTANERA" in s:
        alias.extend(["COSTANERA", "COSTANERA CENTER"])
    if "EGANA" in s or "EGAÑA" in s:
        alias.extend(["PLAZA EGANA", "PLAZA EGAÑA"])
    return list(set(alias))

@costeo_bp.route("/mapeo")
@login_requerido
@permiso_modulo("contab")
def mapeo():
    # 1. Extraer lista única de Productos desde las Ventas
    df_ventas = obtener_datos("comercial")
    if not df_ventas.empty and "DESCRIPCION" in df_ventas.columns:
        productos = sorted(df_ventas["DESCRIPCION"].dropna().unique().tolist())
    else:
        productos = []

    # 2. Extraer solo las Cuentas de Ingresos (Comienzan con 4) desde Contabilidad
    df_mayor = obtener_datos("mayor")
    cuentas_ingreso = []
    if not df_mayor.empty and "CUENTA" in df_mayor.columns:
        df_mayor["CUENTA"] = df_mayor["CUENTA"].astype(str).str.strip()
        df_mayor["NOMBRE"] = df_mayor["NOMBRE"].astype(str).str.strip()
        
        df_ingresos = df_mayor[df_mayor["CUENTA"].str.startswith("4")].drop_duplicates(subset=["CUENTA"])
        for _, row in df_ingresos.iterrows():
            cuentas_ingreso.append({
                "codigo": row["CUENTA"],
                "nombre": row["NOMBRE"],
                "display": f"{row['CUENTA']} - {row['NOMBRE']}"
            })
        cuentas_ingreso = sorted(cuentas_ingreso, key=lambda x: x["codigo"])

    # 3. Traer el mapeo que ya existe en el JSON
    reglas = cargar_reglas()
    mapeo_actual = reglas.get("mapeo_cuentas", {})

    return render_template("contab/costeo_mapeo.html", productos=productos, cuentas=cuentas_ingreso, mapeo_actual=mapeo_actual)

@costeo_bp.route("/api/guardar_mapeo", methods=["POST"])
@login_requerido
@permiso_modulo("contab")
def api_guardar_mapeo():
    data = request.get_json()
    producto, cuenta = data.get("producto"), data.get("cuenta")
    guardar_mapeo(producto, cuenta)
    return jsonify({"success": True})

@costeo_bp.route("/costos_directos")
@login_requerido
@permiso_modulo("contab")
def costos_directos():
    periodo = request.args.get("periodo", datetime.now().strftime("%Y-%m"))
    
    # Extraer lista única de Productos
    df_ventas = obtener_datos("comercial")
    if not df_ventas.empty and "DESCRIPCION" in df_ventas.columns:
        productos = sorted(df_ventas["DESCRIPCION"].dropna().unique().tolist())
    else:
        productos = []

    # Obtener los costos vigentes para el periodo seleccionado
    costos_efectivos, costos_propios = obtener_costos_efectivos(periodo)

    reglas = cargar_reglas()
    historico_bruto = reglas.get("costos_directos_base", {})
    historico_limpio = {}
    for p, hist in historico_bruto.items():
        if isinstance(hist, dict):
            historico_limpio[p] = dict(sorted(hist.items(), reverse=True))
        else:
            historico_limpio[p] = {"2000-01": float(hist)}

    return render_template("contab/costeo_directos.html", 
                           productos=productos, 
                           costos_actuales=costos_efectivos,
                           costos_propios=costos_propios,
                           periodo=periodo,
                           historico=historico_limpio)

@costeo_bp.route("/api/guardar_costo", methods=["POST"])
@login_requerido
@permiso_modulo("contab")
def api_guardar_costo():
    data = request.get_json()
    guardar_costo_directo(data.get("producto"), data.get("costo"), data.get("periodo"))
    return jsonify({"success": True})

@costeo_bp.route("/reglas")
@login_requerido
@permiso_modulo("contab")
def reglas_gastos():
    # Extraer lista de Productos
    df_ventas = obtener_datos("comercial")
    if not df_ventas.empty and "DESCRIPCION" in df_ventas.columns:
        productos = sorted(df_ventas["DESCRIPCION"].dropna().unique().tolist())
    else:
        productos = []

    # Extraer Cuentas de Gasto/Pérdida (Comienzan con 3 o la cuenta contable de gastos)
    df_mayor = obtener_datos("mayor")
    cuentas_gasto = []
    if not df_mayor.empty and "CUENTA" in df_mayor.columns:
        df_mayor["CUENTA"] = df_mayor["CUENTA"].astype(str).str.strip()
        df_mayor["NOMBRE"] = df_mayor["NOMBRE"].astype(str).str.strip()
        
        # NOTA: Asegúrate de que las cuentas de gastos en tu balance empiecen con 3
        df_gastos = df_mayor[df_mayor["CUENTA"].str.startswith("3")].drop_duplicates(subset=["CUENTA"])
        for _, row in df_gastos.iterrows():
            cuentas_gasto.append({
                "codigo": row["CUENTA"],
                "nombre": row["NOMBRE"],
                "display": f"{row['CUENTA']} - {row['NOMBRE']}"
            })
        cuentas_gasto = sorted(cuentas_gasto, key=lambda x: x["codigo"])

    reglas = cargar_reglas()
    reglas_actuales = reglas.get("reglas_gastos", {})

    return render_template("contab/costeo_reglas.html", productos=productos, cuentas=cuentas_gasto, reglas_actuales=reglas_actuales)

@costeo_bp.route("/api/guardar_regla", methods=["POST"])
@login_requerido
@permiso_modulo("contab")
def api_guardar_regla():
    data = request.get_json()
    guardar_regla_gasto(data.get("cuenta"), data.get("regla"))
    return jsonify({"success": True})

@costeo_bp.route("/gav")
@login_requerido
@permiso_modulo("contab")
def gav_corporativo():
    periodo = request.args.get("periodo", datetime.now().strftime("%Y-%m"))
    df_ventas = obtener_datos("comercial")
    sucursales = sorted(df_ventas["SUCURSAL"].dropna().unique().tolist()) if not df_ventas.empty else []
    
    prorrateo_adm = obtener_prorrateo_adm()
    
    prorrateos_data = cargar_prorrateos()
    
    # 1. Obtener regla de SG con herencia histórica
    pool_sg = {}
    for p_key, d in prorrateos_data.get("reglas_mensuales", {}).items():
        if "serv_generales" in d: pool_sg[p_key] = d["serv_generales"]
    ants = [p_key for p_key in pool_sg.keys() if p_key <= periodo]
    regla_efectiva_sg = pool_sg[max(ants)] if ants else {}
    hay_reglas_sg = periodo in pool_sg
    mes_heredado_sg = max(ants) if ants else None

    # 2. Traer Gastos del Periodo desde el Mayor
    df_mayor_raw = obtener_datos("mayor")
    gastos_adm = {}
    gastos_sg = {}
    total_adm = 0.0
    total_sg = 0.0
    asignado_sg_por_sucursal = {s: 0.0 for s in sucursales}

    if not df_mayor_raw.empty and "FECHA" in df_mayor_raw.columns:
        df_mayor_raw["FECHA_DT"] = pd.to_datetime(df_mayor_raw["FECHA"], errors="coerce")
        df_mayor_raw["PERIODO_STR"] = df_mayor_raw["FECHA_DT"].dt.strftime("%Y-%m")
        df_mes = df_mayor_raw[df_mayor_raw["PERIODO_STR"] == periodo].copy()
        
        if not df_mes.empty:
            df_mes["DEBE"] = pd.to_numeric(df_mes["DEBE"], errors="coerce").fillna(0)
            df_mes["HABER"] = pd.to_numeric(df_mes["HABER"], errors="coerce").fillna(0)
            df_mes["SALDO_REAL"] = (df_mes["DEBE"] - df_mes["HABER"]) * -1
            df_mes["CENTRO COSTO"] = df_mes["CENTRO COSTO"].astype(str).str.strip().str.upper()
            df_mes["CUENTA"] = df_mes["CUENTA"].astype(str).str.strip()
            df_mes["NOMBRE"] = df_mes["NOMBRE"].astype(str).str.strip()

            mask_gastos = df_mes["CUENTA"].str.startswith("3")
            df_gastos = df_mes[mask_gastos]

            # Administración
            mask_adm = df_gastos["CENTRO COSTO"].str.contains("ADMINISTRACION", na=False)
            for _, row in df_gastos[mask_adm].iterrows():
                saldo = abs(float(row["SALDO_REAL"]))
                if saldo > 0:
                    cta_display = f"{row['CUENTA']} - {row['NOMBRE']}"
                    gastos_adm[cta_display] = gastos_adm.get(cta_display, 0) + saldo
                    total_adm += saldo

            # Servicios Generales
            mask_sg = df_gastos["CENTRO COSTO"].str.contains("SERVICIOS GENERALES", na=False)
            for _, row in df_gastos[mask_sg].iterrows():
                saldo = abs(float(row["SALDO_REAL"]))
                if saldo > 0:
                    cta_display = f"{row['CUENTA']} - {row['NOMBRE']}"
                    cta_nombre = str(row['NOMBRE'])
                    gastos_sg[cta_display] = gastos_sg.get(cta_display, 0) + saldo
                    total_sg += saldo
                    
                    # Calcular cuánto se le asignó a cada sucursal según las reglas de SG
                    regla_cta = regla_efectiva_sg.get(cta_nombre, {})
                    for suc in sucursales:
                        pct = 0.0
                        aliases_suc = obtener_alias_sucursal(suc)
                        for branch_key, val in regla_cta.items():
                            b_search = branch_key.upper().strip().replace("Ñ", "N")
                            if any(a in b_search or b_search in a for a in aliases_suc):
                                pct = float(val)
                                break
                        asignado_sg_por_sucursal[suc] += saldo * pct

    return render_template("contab/costeo_gav.html", 
                           sucursales=sucursales, 
                           prorrateo_adm=prorrateo_adm, 
                           periodo=periodo, 
                           hay_reglas_sg=hay_reglas_sg,
                           mes_heredado_sg=mes_heredado_sg,
                           gastos_adm=gastos_adm,
                           total_adm=total_adm,
                           gastos_sg=gastos_sg,
                           total_sg=total_sg,
                           asignado_sg_por_sucursal=asignado_sg_por_sucursal)

@costeo_bp.route("/api/guardar_gav_adm", methods=["POST"])
@login_requerido
@permiso_modulo("contab")
def api_guardar_gav_adm():
    data = request.get_json()
    guardar_prorrateo_adm(data.get("distribucion", {}))
    return jsonify({"success": True})

def correr_motor_costeo(periodo, sucursal):
    """Motor matemático centralizado. Devuelve la rentabilidad calculada lista para usar."""
    # 1. Obtener lista de sucursales
    df_ventas_raw = obtener_datos("comercial")
    sucursales = []
    if not df_ventas_raw.empty and "SUCURSAL" in df_ventas_raw.columns:
        sucursales = sorted(df_ventas_raw["SUCURSAL"].dropna().unique().tolist())
        if not sucursal and sucursales:
            sucursal = sucursales[0]

    # 2. Filtrar ventas
    df_ventas = pd.DataFrame()
    if not df_ventas_raw.empty and "FECHA" in df_ventas_raw.columns:
        df_ventas_raw["FECHA_DT"] = pd.to_datetime(df_ventas_raw["FECHA"], errors="coerce")
        df_ventas_raw["PERIODO"] = df_ventas_raw["FECHA_DT"].dt.strftime("%Y-%m")
        df_ventas = df_ventas_raw[(df_ventas_raw["PERIODO"] == periodo) & (df_ventas_raw["SUCURSAL"] == sucursal)].copy()

    ventas_prod = {}
    if not df_ventas.empty:
        for _, row in df_ventas.groupby("DESCRIPCION").agg({"NETO": "sum", "CANTIDAD": "sum"}).reset_index().iterrows():
            if row["CANTIDAD"] > 0:
                ventas_prod[row["DESCRIPCION"]] = {
                    "ingreso": float(row["NETO"]),
                    "unidades": float(row["CANTIDAD"]),
                    "gasto_asignado": 0.0,
                        "desglose_gastos": {},
                        "gav_asignado": 0.0,
                        "desglose_gav": {}
                }

    # 3. Filtrar gastos (AHORA CON MOTOR GERENCIAL DE PRORRATEOS INCLUIDO)
    df_mayor_raw = obtener_datos("mayor")
    gastos_cuenta = {}
    data_config = {}
    df_procesado = pd.DataFrame()
    if not df_mayor_raw.empty and "FECHA" in df_mayor_raw.columns:
        df_mayor_raw["FECHA_DT"] = pd.to_datetime(df_mayor_raw["FECHA"], errors="coerce")
        df_mayor_raw["PERIODO_STR"] = df_mayor_raw["FECHA_DT"].dt.strftime("%Y-%m")
        df_mes = df_mayor_raw[df_mayor_raw["PERIODO_STR"] == periodo].copy()
        
        if not df_mes.empty:
            # A. Formatear para el motor de prorrateos
            df_mes["DEBE"] = pd.to_numeric(df_mes["DEBE"], errors="coerce").fillna(0)
            df_mes["HABER"] = pd.to_numeric(df_mes["HABER"], errors="coerce").fillna(0)
            df_mes["SALDO_REAL"] = (df_mes["DEBE"] - df_mes["HABER"]) * -1
            df_mes["CENTRO COSTO"] = df_mes["CENTRO COSTO"].astype(str).str.strip().str.upper()
            df_mes["CUENTA"] = df_mes["CUENTA"].astype(str).str.strip()
            df_mes["NOMBRE"] = df_mes["NOMBRE"].astype(str).str.strip()
            
            # B. Cargar configuraciones de prorrateo
            prorrateos_data = cargar_prorrateos()
            data_config = {
                "config_cuentas": prorrateos_data.get("config_cuentas", {}),
                "reglas_mensuales": prorrateos_data.get("reglas_mensuales", {}),
                "fabrica_empanadas": prorrateos_data.get("fabrica_empanadas", {})
            }
            
            # C. Juguera Gerencial: Apagamos SG para aislarlo, dejamos Fábrica encendida
            df_procesado = calcular_matriz_gestion(df_mes, periodo, switch_sg=False, switch_fab=True, data_config=data_config)
            
            # D. Extraer solo los gastos (3) propios de la sucursal (Búsqueda Flexible)
            df_gastos = pd.DataFrame()
            if not df_procesado.empty and "CUENTA" in df_procesado.columns and "CENTRO COSTO" in df_procesado.columns:
                aliases_suc = obtener_alias_sucursal(sucursal)
                cc_normalized = df_procesado["CENTRO COSTO"].astype(str).str.upper().str.strip().str.replace("Ñ", "N", regex=False)
                mask_cc = cc_normalized.apply(lambda x: any((a in x or x in a) for a in aliases_suc) if str(x).strip() else False)
                mask = (df_procesado["CUENTA"].str.startswith("3")) & mask_cc
                df_gastos = df_procesado[mask].copy()
            
            if not df_gastos.empty:
                df_gastos["DISPLAY"] = df_gastos["CUENTA"] + " - " + df_gastos["NOMBRE"]
                
                for _, row in df_gastos.groupby("DISPLAY")["SALDO_REAL"].sum().reset_index().iterrows():
                    # Los gastos en SALDO_REAL son negativos. Pasamos a absoluto.
                    saldo_positivo = abs(float(row["SALDO_REAL"]))
                    if saldo_positivo > 0:
                        gastos_cuenta[row["DISPLAY"]] = saldo_positivo

    # 4. Aplicar reglas
    reglas = cargar_reglas()
    reglas_gastos = reglas.get("reglas_gastos", {})
    costos_directos, _ = obtener_costos_efectivos(periodo)

    for cta_display, monto_gasto in gastos_cuenta.items():
        regla = reglas_gastos.get(cta_display)
        
        # Fallback inteligente: Si el motor renombró la cuenta, buscar por código numérico
        if not regla:
            codigo_cuenta = str(cta_display).split(" - ")[0].strip()
            for key_json, rule_json in reglas_gastos.items():
                if str(key_json).startswith(codigo_cuenta + " -"):
                    regla = rule_json
                    break
                    
        if not regla:
            continue
        
        alcance = regla.get("alcance", "global")
        metodo = regla.get("metodo", "venta_dinero")
        afectados = regla.get("productos_afectados", []) if alcance == "especifico" else list(ventas_prod.keys())
        
        prods_validos = [p for p in afectados if p in ventas_prod and ventas_prod[p]["unidades"] > 0]
        if not prods_validos:
            continue
            
        if metodo == "venta_dinero":
            base_total = sum(ventas_prod[p]["ingreso"] for p in prods_validos)
        else:
            base_total = sum(ventas_prod[p]["unidades"] for p in prods_validos)
            
        if base_total == 0:
            continue
            
        for p in prods_validos:
            base_prod = ventas_prod[p]["ingreso"] if metodo == "venta_dinero" else ventas_prod[p]["unidades"]
            proporcion = base_prod / base_total
            asignado = monto_gasto * proporcion
            ventas_prod[p]["gasto_asignado"] += asignado
            ventas_prod[p]["desglose_gastos"][cta_display] = asignado

    # ========================================================
    # 5. NUEVO CÁLCULO GAV (SG + ADMINISTRACIÓN)
    # ========================================================
    total_gav_sucursal = 0.0
    desglose_gav_total = {}
    
    # A. SG (Histórico desde el motor gerencial)
    pool_sg = {}
    for p_key, d in data_config.get("reglas_mensuales", {}).items():
        if "serv_generales" in d: pool_sg[p_key] = d["serv_generales"]
    ants = [p_key for p_key in pool_sg.keys() if p_key <= periodo]
    regla_efectiva_sg = pool_sg[max(ants)] if ants else {}
    
    if not df_procesado.empty and "CUENTA" in df_procesado.columns and "CENTRO COSTO" in df_procesado.columns:
        mask_sg = (df_procesado["CUENTA"].str.startswith("3")) & (df_procesado["CENTRO COSTO"].str.contains("SERVICIOS GENERALES", na=False))
        for _, row in df_procesado[mask_sg].iterrows():
            cta_nombre = str(row["NOMBRE"])
            monto = abs(float(row["SALDO_REAL"]))
            pct = 0.0
            aliases_suc = obtener_alias_sucursal(sucursal)
            for branch_key, val in regla_efectiva_sg.get(cta_nombre, {}).items():
                b_search = branch_key.upper().strip().replace("Ñ", "N")
                if any(a in b_search or b_search in a for a in aliases_suc):
                    pct = float(val)
                    break
            if pct > 0:
                asignado = monto * pct
                total_gav_sucursal += asignado
                desglose_gav_total[f"{row['CUENTA']} - {row['NOMBRE']} (SS.GG)"] = asignado

        # B. Administración (Desde la nueva pestaña)
        prorrateo_adm = obtener_prorrateo_adm()
        pct_adm = float(prorrateo_adm.get(sucursal, 0)) / 100.0
        mask_adm = (df_procesado["CUENTA"].str.startswith("3")) & (df_procesado["CENTRO COSTO"].str.contains("ADMINISTRACION", na=False))
        for _, row in df_procesado[mask_adm].iterrows():
            monto = abs(float(row["SALDO_REAL"]))
            if pct_adm > 0 and monto > 0:
                asignado = monto * pct_adm
                total_gav_sucursal += asignado
                desglose_gav_total[f"{row['CUENTA']} - {row['NOMBRE']} (ADM)"] = asignado

    # C. Distribuir el GAV total en base al Ingreso ($)
    total_ingreso_sucursal = sum(p["ingreso"] for p in ventas_prod.values())
    
    for p, data in ventas_prod.items():
        proporcion_ingreso = data["ingreso"] / total_ingreso_sucursal if total_ingreso_sucursal > 0 else 0
        data["gav_asignado"] = total_gav_sucursal * proporcion_ingreso
        data["desglose_gav"] = {k: v * proporcion_ingreso for k, v in desglose_gav_total.items()}

    # 5. Formatear vista
    resultados = []
    total_gasto_sucursal = sum(gastos_cuenta.values())
    gasto_asignado_total = sum(p["gasto_asignado"] for p in ventas_prod.values())

    for prod, data in ventas_prod.items():
        uni = data["unidades"]
        ing = data["ingreso"]
        precio_prom = ing / uni
        c_dir_uni = float(costos_directos.get(prod, 0))
        c_dir_tot = c_dir_uni * uni
        g_asig_tot = data["gasto_asignado"]
        g_asig_uni = g_asig_tot / uni
        gav_tot = data["gav_asignado"]
        gav_uni = gav_tot / uni
        c_tot_uni = c_dir_uni + g_asig_uni + gav_uni
        margen_uni = precio_prom - c_tot_uni
        margen_pct = (margen_uni / precio_prom * 100) if precio_prom > 0 else 0
        
        resultados.append({
            "producto": prod,
            "unidades": int(uni),
            "ingreso": ing,
            "precio_prom": precio_prom,
            "costo_directo_uni": c_dir_uni,
            "gasto_fijo_uni": g_asig_uni,
            "gasto_fijo_tot": g_asig_tot,
            "gav_uni": gav_uni,
            "gav_tot": gav_tot,
            "costo_total_uni": c_tot_uni,
            "margen_uni": margen_uni,
            "margen_pct": margen_pct,
            "desglose": data["desglose_gastos"],
            "desglose_gav": data["desglose_gav"]
        })

    resultados.sort(key=lambda x: x["ingreso"], reverse=True)

    return sucursales, sucursal, resultados, total_ingreso_sucursal, total_gasto_sucursal, gasto_asignado_total, gastos_cuenta

@costeo_bp.route("/simulador")
@login_requerido
@permiso_modulo("contab")
def simulador():
    periodo = request.args.get("periodo", datetime.now().strftime("%Y-%m"))
    sucursal = request.args.get("sucursal", "")
    
    sucursales, sucursal, resultados, total_ingreso_sucursal, total_gasto_sucursal, gasto_asignado_total, gastos_cuenta = correr_motor_costeo(periodo, sucursal)

    return render_template("contab/costeo_simulador.html",
                           periodo=periodo,
                           sucursal=sucursal,
                           sucursales=sucursales,
                           resultados=resultados,
                           total_ingreso=total_ingreso_sucursal,
                           total_gasto=total_gasto_sucursal,
                           gasto_asignado=gasto_asignado_total,
                           gastos_cuenta=gastos_cuenta)

@costeo_bp.route("/exportar_simulador")
@login_requerido
@permiso_modulo("contab")
def exportar_simulador():
    periodo = request.args.get("periodo", datetime.now().strftime("%Y-%m"))
    sucursal = request.args.get("sucursal", "")
    
    _, _, resultados, _, _, _, _ = correr_motor_costeo(periodo, sucursal)
    
    df_exp = pd.DataFrame(resultados)
    if df_exp.empty:
        return "No hay datos para exportar", 204
        
    df_exp = df_exp[["producto", "unidades", "ingreso", "precio_prom", "costo_directo_uni", "gasto_fijo_uni", "gav_uni", "costo_total_uni", "margen_uni", "margen_pct"]]
    df_exp.columns = ["Producto", "Unidades", "Ingreso Total", "Precio Prom.", "Costo Directo Uni.", "Gasto Fijo Local Uni.", "GAV Corporativo Uni.", "Costo Total Uni.", "Margen Neto Uni.", "Margen Neto %"]
    
    output = io.BytesIO()
    df_exp.to_excel(output, index=False)
    output.seek(0)
    
    wb = load_workbook(output)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=3):
        for cell in row: aplicar_formato_numerico_excel(cell)
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name=f"Costeo_{sucursal}_{periodo}.xlsx", as_attachment=True)

@costeo_bp.route("/rentabilidad_gerencia")
@login_requerido
@permiso_modulo("reporte")
def rentabilidad_gerencia():
    periodo = request.args.get("periodo", datetime.now().strftime("%Y-%m"))
    sucursal = request.args.get("sucursal", "")
    
    sucursales, sucursal, resultados, total_ingreso, _, _, _ = correr_motor_costeo(periodo, sucursal)
    
    costo_total_op = sum(r["costo_total_uni"] * r["unidades"] for r in resultados)
    margen_neto = total_ingreso - costo_total_op
    margen_pct = (margen_neto / total_ingreso * 100) if total_ingreso > 0 else 0
    
    top_resultados = sorted(resultados, key=lambda x: x["margen_uni"] * x["unidades"], reverse=True)[:10]
    nombres_grafico = [r["producto"] for r in top_resultados]
    margenes_grafico = [r["margen_uni"] * r["unidades"] for r in top_resultados]
    
    return render_template("contab/rentabilidad_gerencia.html",
                           periodo=periodo,
                           sucursal=sucursal,
                           sucursales=sucursales,
                           resultados=resultados,
                           total_ingreso=total_ingreso,
                           costo_total_op=costo_total_op,
                           margen_neto=margen_neto,
                           margen_pct=margen_pct,
                           nombres_grafico=nombres_grafico,
                           margenes_grafico=margenes_grafico)