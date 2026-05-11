from flask import Blueprint, render_template, request, jsonify
from utils.sheet_cache import obtener_datos
from utils.filters import filtrar_dataframe
from services.resumen_service import obtener_resumen_mensual_tabular, MESES
from services.detalle_service import obtener_detalle
import pandas as pd
from utils.utils_excel import aplicar_formato_numerico_excel
from openpyxl import load_workbook
import io
from flask import send_file
from utils.sheet_cache import obtener_fecha_actualizacion
from utils.auth import login_requerido, permiso_modulo
from services.ventas_por_dia_service import obtener_detalle_por_dia






ventas_bp = Blueprint("ventas", __name__)

#@ventas_bp.route("/ventas", methods=["GET"])
@ventas_bp.route("/ventas", methods=["GET"])
@login_requerido
@permiso_modulo("ventas")
def ventas():
    empresa = request.args.get("empresa", "comercial")
    df = obtener_datos(empresa)

    # Obtener campos del formulario
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    año_actual = request.args.get("año")
    semana_actual = request.args.get("semana")

    # Sanitizar explícitamente strings vacíos para evitar anulación de filtros
    if desde == "": desde = None
    if hasta == "": hasta = None
    if año_actual == "": año_actual = None
    if semana_actual == "": semana_actual = None

    # Si no hay fechas, aplicar semana automática solo si no se recibe por GET
    if not desde and not hasta:
        if not semana_actual:
            if año_actual:
                if not df.empty and "AÑO" in df.columns and "SEMANA" in df.columns:
                    semana_actual = df[df["AÑO"] == int(año_actual)]["SEMANA"].max()
            else:
                if not df.empty and "AÑO" in df.columns and "SEMANA" in df.columns:
                    año_actual = df["AÑO"].max()
                    semana_actual = df[df["AÑO"] == año_actual]["SEMANA"].max()
        else:
            semana_actual = int(semana_actual)
    else:
        # Si hay fechas, ignorar semana/año
        semana_actual = None
        año_actual = None

    # Filtros
    filtros = {
        "empresa": empresa,
        "sucursal": request.args.get("sucursal"),
        "semana": semana_actual,
        "año": año_actual,
        "desde": desde,
        "hasta": hasta,
        "filtro_por": request.args.get("filtro_por", "FAMILIA"),
        "valor": request.args.get("valor", "TODOS"),
        "tab": request.args.get("tab", "detalle")  # << ESTA LÍNEA es CLAVE
    }


    # Aplica lógica de control para evitar conflicto entre fechas y semana
    if filtros["desde"] or filtros["hasta"]:
        semana = None
        año = None
    else:
        semana = filtros["semana"]
        año = filtros["año"]

    detalle_df = filtrar_dataframe(
        df,
        filtros["filtro_por"],
        filtros["valor"],
        filtros["sucursal"],
        semana,
        año,
        filtros["desde"],
        filtros["hasta"]
    )

    # === Lógica para la nueva pestaña "Detalle por Día" ===
    filtros["dia_semana"] = request.args.get("dia_semana", "TODOS")

    # El servicio interno se encargará de filtrar por día de la semana.
    if not detalle_df.empty:
        detalle_dia_data = obtener_detalle_por_dia(detalle_df.copy(), filtros)
    else:
        detalle_dia_data = {"tabla": [], "columnas": [], "titulo": "No hay datos para los filtros seleccionados"}

    # Generar detalle
# Generar detalle agrupado visualmente (sin afectar Excel)
    if isinstance(detalle_df, pd.DataFrame) and not detalle_df.empty:
        # Agrupamos dinámicamente según el filtro seleccionado
        campo_agrupacion = filtros.get("filtro_por", "FAMILIA")
        if campo_agrupacion not in detalle_df.columns:
            campo_agrupacion = "DESCRIPCION"  # Fallback de seguridad
            
        detalle_agrupado = (
            detalle_df.groupby(campo_agrupacion, as_index=False)
            .agg(
                CANTIDAD=("CANTIDAD", "sum"),
                NETO=("NETO", "sum")
            )
        ).rename(columns={campo_agrupacion: "PRODUCTO"}) # Renombramos para la vista

        # Calculamos el precio unitario de forma segura
        detalle_agrupado["PRECIO_UNITARIO"] = detalle_agrupado.apply(
            lambda row: row["NETO"] / row["CANTIDAD"] if row["CANTIDAD"] != 0 else 0,
            axis=1
        )
        detalle = detalle_agrupado.to_dict(orient="records")
    else:
        detalle = []

    # Generar resumen mensual
    resumen_mensual = obtener_resumen_mensual_tabular(df, filtros)


    # Select dinámico para valores
    filtro_por = filtros["filtro_por"]
    opciones_valor = []
    if filtro_por in df.columns:
        opciones_valor = sorted(df[filtro_por].dropna().unique())

    # Lista de sucursales
    sucursales = []
    if "SUCURSAL" in df.columns:
        sucursales = sorted(df["SUCURSAL"].dropna().unique().tolist())
    sucursales.insert(0, "TODAS")

    return render_template("ventas.html",
        detalle=detalle,
        resumen_mensual=resumen_mensual,
        meses=MESES,
        filtros=filtros,
        opciones_valor=opciones_valor,
        filtro_por=filtro_por,
        filtro_valor=filtros["valor"],
        sucursal=filtros["sucursal"],
        empresa=empresa,
        sucursales=sucursales,
        fecha_actualizacion=obtener_fecha_actualizacion("temperatura_equipos"),
        detalle_dia_data=detalle_dia_data)



@ventas_bp.route("/descargar_excel")
@login_requerido
def descargar_excel():
    empresa = request.args.get("empresa", "comercial")
    df = obtener_datos(empresa)

    sucursal = request.args.get("sucursal")
    semana = request.args.get("semana")
    año = request.args.get("año")
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    filtro_por = request.args.get("filtro_por", "FAMILIA")
    valor = request.args.get("valor", "TODOS")
    tab = request.args.get("tab", "detalle")

    # Sanitizar explícitamente
    if desde == "": desde = None
    if hasta == "": hasta = None
    if año == "": año = None
    if semana == "": semana = None

    filtros = {
        "empresa": empresa,
        "sucursal": sucursal,
        "semana": semana,
        "año": año,
        "desde": desde,
        "hasta": hasta,
        "filtro_por": filtro_por,
        "valor": valor
    }

    output = io.BytesIO()

    if tab == "detalle":
        df_filtrado = filtrar_dataframe(df, filtro_por, valor, sucursal, semana, año, desde, hasta)
        if df_filtrado.empty:
            return "No hay datos para exportar", 204

        df_filtrado.to_excel(output, index=False)
        output.seek(0)

        # Formatear usando openpyxl
        wb = load_workbook(output)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, min_col=1):
            for cell in row:
                aplicar_formato_numerico_excel(cell)

        output = io.BytesIO()
        wb.save(output)

    elif tab == "resumen":
        resumen = obtener_resumen_mensual_tabular(df, filtros)
        if not resumen or not resumen.get("tabla"):
            return "No hay resumen para exportar", 204

        filas = []
        for item in resumen["tabla"]:
            neto_f = [float(str(v).replace("$", "").replace(".", "").replace(",", ".").strip()) for v in item["neto"]]
            unit_f = [float(str(v).replace("$", "").replace(".", "").replace(",", ".").strip()) for v in item["unit"]]
            cant_f = [int(str(v).replace(".", "").strip()) for v in item["cant"]]

            filas.append(["", item["producto"]] + neto_f + [float(item["total_neto"].replace("$", "").replace(".", "").replace(",", ".").strip())])
            filas.append(["", ""] + cant_f + [int(item["total_cant"].replace(".", "").strip())])
            filas.append(["", ""] + unit_f + [float(item["total_unit"].replace("$", "").replace(".", "").replace(",", ".").strip())])

        encabezados = ["Tipo", "Producto"] + MESES + ["Total"]
        df_export = pd.DataFrame(filas, columns=encabezados)
        df_export.to_excel(output, index=False)
        output.seek(0)

        wb = load_workbook(output)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, min_col=3):  # Desde col C (primer número)
            for cell in row:
                aplicar_formato_numerico_excel(cell)

        output = io.BytesIO()
        wb.save(output)

    output.seek(0)
    return send_file(output,
                     download_name=f"ventas_{tab}.xlsx",
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# =====================================================================
# VISTA HISTÓRICO DE PRODUCTOS
# =====================================================================

@ventas_bp.route("/ventas/historico")
@login_requerido
@permiso_modulo("ventas")
def ventas_historico():
    return render_template("ventas_historico.html")


@ventas_bp.route("/api/historico-resumen")
@login_requerido
def api_historico_resumen():
    """Cards de productos significativos: top neto, mayor crecimiento, mayor caída."""
    empresa = request.args.get("empresa", "comercial")
    sucursal = request.args.get("sucursal")
    familia = request.args.get("familia")

    df = obtener_datos(empresa)
    if df.empty:
        return jsonify({"top_neto": [], "top_crecimiento": [], "top_caida": [],
                        "sucursales": [], "familias": []})

    sucursales = sorted(df["SUCURSAL"].dropna().unique().tolist()) if "SUCURSAL" in df.columns else []
    familias = sorted(df["FAMILIA"].dropna().unique().tolist()) if "FAMILIA" in df.columns else []

    df_f = filtrar_dataframe(df, tipo="FAMILIA", valor=familia or "TODOS",
                             sucursal=sucursal, semana=None, año=None, desde=None, hasta=None)
    if df_f.empty:
        return jsonify({"top_neto": [], "top_crecimiento": [], "top_caida": [],
                        "sucursales": sucursales, "familias": familias})

    año_max = int(df_f["AÑO"].max())
    año_ant = año_max - 1

    curr = df_f[df_f["AÑO"] == año_max].groupby("DESCRIPCION").agg(
        neto=("NETO", "sum"), cantidad=("CANTIDAD", "sum")).reset_index()
    prev = df_f[df_f["AÑO"] == año_ant].groupby("DESCRIPCION").agg(
        neto=("NETO", "sum"), cantidad=("CANTIDAD", "sum")).reset_index()

    merged = pd.merge(curr, prev, on="DESCRIPCION", how="outer", suffixes=("_act", "_ant")).fillna(0)
    merged["variacion"] = merged["neto_act"] - merged["neto_ant"]
    merged["var_pct"] = merged.apply(
        lambda r: round((r["variacion"] / abs(r["neto_ant"])) * 100, 1) if abs(r["neto_ant"]) > 0 else 0, axis=1)

    # Sparkline: neto semanal del año actual (últimas 12 semanas con datos)
    semanas_disp = sorted(df_f[df_f["AÑO"] == año_max]["SEMANA"].dropna().unique())
    ultimas_sem = semanas_disp[-12:] if len(semanas_disp) > 12 else semanas_disp
    spark_base = df_f[(df_f["AÑO"] == año_max) & (df_f["SEMANA"].isin(ultimas_sem))]
    spark_data = spark_base.groupby(["DESCRIPCION", "SEMANA"])["NETO"].sum().reset_index()

    def build_spark(desc):
        s = spark_data[spark_data["DESCRIPCION"] == desc].sort_values("SEMANA")
        return s["NETO"].tolist()

    def to_cards(subset, n=5):
        cards = []
        for _, r in subset.head(n).iterrows():
            cards.append({
                "producto": str(r["DESCRIPCION"]),
                "neto_actual": int(r["neto_act"]),
                "neto_anterior": int(r["neto_ant"]),
                "variacion": int(r["variacion"]),
                "var_pct": float(r["var_pct"]),
                "cantidad_actual": int(r["cantidad_act"]),
                "sparkline": build_spark(r["DESCRIPCION"])
            })
        return cards

    top_neto = to_cards(merged.sort_values("neto_act", ascending=False))
    top_crec = to_cards(merged[merged["variacion"] > 0].sort_values("variacion", ascending=False))
    top_caida = to_cards(merged[merged["variacion"] < 0].sort_values("variacion", ascending=True))

    productos_lista = sorted(df_f["DESCRIPCION"].dropna().unique().tolist())

    return jsonify({
        "top_neto": top_neto,
        "top_crecimiento": top_crec,
        "top_caida": top_caida,
        "sucursales": sucursales,
        "familias": familias,
        "productos": productos_lista,
        "año_actual": año_max,
        "año_anterior": año_ant,
    })


@ventas_bp.route("/api/historico-producto")
@login_requerido
def api_historico_producto():
    """Detalle histórico de un producto: series semanales de neto, cantidad y precio."""
    empresa = request.args.get("empresa", "comercial")
    sucursal = request.args.get("sucursal")
    familia = request.args.get("familia")
    producto = request.args.get("producto")

    if not producto:
        return jsonify({"error": "Falta parámetro producto"}), 400

    df = obtener_datos(empresa)
    df_f = filtrar_dataframe(df, tipo="FAMILIA", valor=familia or "TODOS",
                             sucursal=sucursal, semana=None, año=None, desde=None, hasta=None)

    df_prod = df_f[df_f["DESCRIPCION"] == producto].copy()
    if df_prod.empty:
        return jsonify({"semanas": [], "neto_actual": [], "neto_anterior": [],
                        "cant_actual": [], "cant_anterior": [],
                        "precio_actual": [], "precio_anterior": [], "resumen_mensual": []})

    año_max = int(df_prod["AÑO"].max())
    año_ant = año_max - 1

    def serie_semanal(df_año, año):
        d = df_año[df_año["AÑO"] == año]
        grp = d.groupby("SEMANA").agg(neto=("NETO", "sum"), cantidad=("CANTIDAD", "sum")).reset_index()
        grp["precio"] = grp.apply(lambda r: round(r["neto"] / r["cantidad"]) if r["cantidad"] != 0 else 0, axis=1)
        return grp.sort_values("SEMANA")

    act = serie_semanal(df_prod, año_max)
    ant = serie_semanal(df_prod, año_ant)

    all_sems = sorted(set(act["SEMANA"].tolist() + ant["SEMANA"].tolist()))

    act_idx = act.set_index("SEMANA")
    ant_idx = ant.set_index("SEMANA")

    semanas = [int(s) for s in all_sems]
    neto_act = [int(act_idx.loc[s, "neto"]) if s in act_idx.index else 0 for s in all_sems]
    neto_ant = [int(ant_idx.loc[s, "neto"]) if s in ant_idx.index else 0 for s in all_sems]
    cant_act = [int(act_idx.loc[s, "cantidad"]) if s in act_idx.index else 0 for s in all_sems]
    cant_ant = [int(ant_idx.loc[s, "cantidad"]) if s in ant_idx.index else 0 for s in all_sems]
    prec_act = [int(act_idx.loc[s, "precio"]) if s in act_idx.index else 0 for s in all_sems]
    prec_ant = [int(ant_idx.loc[s, "precio"]) if s in ant_idx.index else 0 for s in all_sems]

    # Resumen mensual
    _MESES_ES = {1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
                 7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"}
    df_prod["FECHA_DT"] = pd.to_datetime(df_prod["FECHA"], errors="coerce")
    df_prod["MES"] = df_prod["FECHA_DT"].dt.month

    resumen = []
    for mes in range(1, 13):
        d_act = df_prod[(df_prod["AÑO"] == año_max) & (df_prod["MES"] == mes)]
        d_ant = df_prod[(df_prod["AÑO"] == año_ant) & (df_prod["MES"] == mes)]
        n_act = int(d_act["NETO"].sum())
        n_ant = int(d_ant["NETO"].sum())
        c_act = int(d_act["CANTIDAD"].sum())
        c_ant = int(d_ant["CANTIDAD"].sum())
        resumen.append({
            "mes": _MESES_ES[mes],
            "neto_actual": n_act, "neto_anterior": n_ant,
            "cantidad_actual": c_act, "cantidad_anterior": c_ant,
            "precio_actual": round(n_act / c_act) if c_act else 0,
            "precio_anterior": round(n_ant / c_ant) if c_ant else 0,
        })

    total_neto_act = int(df_prod[df_prod["AÑO"] == año_max]["NETO"].sum())
    total_neto_ant = int(df_prod[df_prod["AÑO"] == año_ant]["NETO"].sum())
    total_cant_act = int(df_prod[df_prod["AÑO"] == año_max]["CANTIDAD"].sum())
    total_cant_ant = int(df_prod[df_prod["AÑO"] == año_ant]["CANTIDAD"].sum())

    return jsonify({
        "producto": producto,
        "año_actual": año_max, "año_anterior": año_ant,
        "semanas": semanas,
        "neto_actual": neto_act, "neto_anterior": neto_ant,
        "cant_actual": cant_act, "cant_anterior": cant_ant,
        "precio_actual": prec_act, "precio_anterior": prec_ant,
        "resumen_mensual": resumen,
        "totales": {
            "neto_actual": total_neto_act, "neto_anterior": total_neto_ant,
            "cantidad_actual": total_cant_act, "cantidad_anterior": total_cant_ant,
            "precio_actual": round(total_neto_act / total_cant_act) if total_cant_act else 0,
            "precio_anterior": round(total_neto_ant / total_cant_ant) if total_cant_ant else 0,
        }
    })
